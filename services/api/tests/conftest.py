import io
import os
import shutil
import time

import httpx
import pytest
from openpyxl import Workbook

from services.api.app.domain.ports import EmbeddingModelPort
from services.api.app.infra.embedding_model import EmbeddingModel

BASE = os.getenv("BASE", "http://localhost:8000")

@pytest.fixture(scope="session")
def tesseract_available():
    return shutil.which("tesseract") is not None

def pytest_collection_modifyitems(config, items):
    # Skip @e2e unless -k e2e is explicitly requested
    k = config.getoption("-k") or ""
    if "e2e" not in k:
        skip_e2e = pytest.mark.skip(reason="Skipping e2e by default; run with -k e2e")
        for item in items:
            if "e2e" in item.keywords:
                item.add_marker(skip_e2e)


@pytest.fixture(scope="session")
def wait_for_api():
    """Wait for /live up to ~60s; skip e2e if not available."""
    if httpx is None:
        pytest.skip("httpx not installed; skipping e2e")
    url = f"{BASE}/live"
    for _ in range(60):
        try:
            r = httpx.get(url, timeout=2.0)
            if r.status_code == 200:
                return
        except Exception:
            pass
        time.sleep(1)
    pytest.skip("API not reachable at /health; skipping e2e")

@pytest.fixture(scope="session")
def ensure_seed(wait_for_api):
    # Create & embed a tiny doc so search has something to find.
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello seed content"
    ws["B2"] = "Cargorisk test"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    files = {
        'file': (
            'seed.xlsx',
            xlsx_bytes,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }

    r = httpx.post(f"{BASE}/upload", files=files, timeout=20)
    r.raise_for_status()
    asset_id = r.json()['id']

    httpx.post(f"{BASE}/document/{asset_id}/ocr", timeout=60).raise_for_status()
    httpx.post(f"{BASE}/document/{asset_id}/embed", timeout=60).raise_for_status()
    return asset_id


@pytest.fixture(scope="session")
def embedding_model() -> EmbeddingModelPort:
    return EmbeddingModel()