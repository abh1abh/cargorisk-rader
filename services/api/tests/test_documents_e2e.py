import io
import os

import httpx
import pytest
from openpyxl import Workbook

BASE = os.getenv("BASE", "http://localhost:8000")

@pytest.mark.e2e
def test_upload_e2e(tesseract_available):
    # create a simple text file
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello seed content"
    ws["B2"] = "Cargorisk test"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    files = {
        'file': (
            'seed.xlsx',
            xlsx_bytes,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }
    r = httpx.post(f"{BASE}/upload", files=files, timeout=30)
    r.raise_for_status()
    asset_id = r.json()['id']

    if not tesseract_available:
        pytest.skip("tesseract not available in PATH")

    r = httpx.post(f"{BASE}/documents/{asset_id}/ocr", timeout=120)
    r.raise_for_status()

    r = httpx.get(f"{BASE}/documents/{asset_id}/text", timeout=30)
    r.raise_for_status()
    assert "Hello" in r.json()['text']
