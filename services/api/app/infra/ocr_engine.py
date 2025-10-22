import io
from dataclasses import dataclass

import fitz  # PyMuPDF
import pytesseract
from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageOps

"""
    Wraps OCR-related operations:
      - PDF: prefer native text; fall back to rasterize+OCR for image-only pages
      - Images: light denoise + contrast, then Tesseract
      - XLSX/XLSM: flatten cells across sheets with an upper bound

    Config:
      lang: Tesseract language
      pdf_dpi: rasterization DPI when OCRing image-only PDF pages
      oem: Tesseract OCR Engine Mode (default 3: default)
      psm: Page segmentation mode (default 6: Assume a single uniform block of text)
    """
@dataclass(slots=True)
class OcrEngine:
    lang: str = "eng+nor"
    pdf_dpi: int = 200
    oem: int = 3
    psm: int = 6

    def pdf_bytes_to_text(self, pdf_bytes: bytes, lang: str | None = None) -> str:
        eff_lang = (lang or self.lang)
        parts: list[str] = []
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                # 1) Try native text (fast & accurate for digital PDFs)
                native = page.get_text("text") or ""
                if native.strip():
                    parts.append(native)
                    continue

                # 2) Rasterize + OCR for image-only pages
                pix = page.get_pixmap(dpi=self.pdf_dpi, alpha=False)
                img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
                parts.append(self._ocr_image(img, lang=eff_lang))
        return "\n".join(parts).strip()

    def image_bytes_to_text(self, img_bytes: bytes, lang: str | None = None) -> str:
        eff_lang = (lang or self.lang)
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        return self._ocr_image(img, lang=eff_lang).strip()

    def xlsx_bytes_to_text(self, xlsx_bytes: bytes, max_cells: int = 2000) -> str:
        """
        Extracts text from an XLSX/XLSM file (all sheets) and flattens to a string.
        Limits cells to avoid huge payloads for embedding/search.
        """
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        parts: list[str] = []
        cells_seen = 0

        for ws in wb.worksheets:
            parts.append(f"\n=== Sheet: {ws.title} ===\n")
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                row_txt = "\t".join("" if v is None else str(v).strip() for v in row)
                if row_txt.strip():
                    parts.append(row_txt + "\n")
                cells_seen += len(row or [])
                if cells_seen >= max_cells:
                    parts.append("\n[TRUNCATED: too many cells]\n")
                    return "".join(parts).strip()

        return "".join(parts).strip()


    def _ocr_image(self, img: Image.Image, lang: str) -> str:
        # light denoise and contrast enhancement
        gray = img.convert("L")
        gray = ImageOps.autocontrast(gray)
        gray = gray.filter(ImageFilter.MedianFilter(size=3))
        return pytesseract.image_to_string(gray, lang=lang, config=self._tesseract_config)


