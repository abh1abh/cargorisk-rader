import io

import fitz  # PyMuPDF
import pytesseract
from PIL import Image, ImageFilter, ImageOps
from openpyxl import load_workbook

def _ocr_image(img: Image.Image, lang: str = "eng+nor") -> str:
    # light denoise and contrast enhancement
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    img = img.filter(ImageFilter.MedianFilter(size=3))
    return pytesseract.image_to_string(img, lang=lang, config="--oem 3 --psm 6")
def pdf_bytes_to_text(pdf_bytes: bytes, lang: str = "eng+nor") -> str:
    text_parts: list[str] = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            # 1) Native text first (fast, accurate for digital PDFs)
            native = page.get_text("text") or ""
            if native.strip():
                text_parts.append(native)
                continue
            # 2) Rasterize & OCR for image-only pages
            pix = page.get_pixmap(dpi=200, alpha=False)   # dpi 200 is a good trade-off
            img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
            text_parts.append(_ocr_image(img, lang=lang))
    return "\n".join(text_parts).strip()

def image_bytes_to_text(img_bytes: bytes, lang: str = "eng+nor") -> str:
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    return _ocr_image(img, lang=lang).strip()

def xlsx_bytes_to_text(xlsx_bytes: bytes, max_cells: int = 20000) -> str:
    """
    Extracts text from an XLSX/XLSM file (all sheets) and flattens to a string.
    Limits the number of cells to avoid huge embeddings.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    parts = []
    cells_seen = 0
    for ws in wb.worksheets:
        parts.append(f"\n=== Sheet: {ws.title} ===\n")
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            row_txt = "\t".join("" if v is None else str(v).strip() for v in row)
            if row_txt.strip():
                parts.append(row_txt + "\n")
            cells_seen += len(row or [])
            if cells_seen >= max_cells:
                parts.append("\n[TRUNCATED: too many cells]\n")
                return "".join(parts).strip()
    return "".join(parts).strip()
